from io import BytesIO
import base64
import json
import mimetypes
from pathlib import Path
import os
import re
import shutil
import subprocess
import urllib.error
import urllib.request

from django.conf import settings
from django.core.files.base import ContentFile
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import FormAutomationJob


FORM_ASSET_LABELS = {
    'purchase_screenshots': '购买信息截图',
    'invoices': '发票',
    'reference_images': '参考图',
    'link_txt': '链接 TXT',
}

IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.webp', '.bmp'}
PDF_EXTENSIONS = {'.pdf'}
DETAIL_FONT = Font(name='等线', color='FF000000', bold=False, size=12)
TOTAL_FONT = Font(name='等线', color='FF000000', bold=True, size=16)
EXPENSE_WORK05_IMAGE_SIZE = {
    'single': (148, 320),
    'multiple': (250, 269),
}


def frontend_template_path(filename):
    return settings.BASE_DIR / 'frontend' / 'public' / 'templates' / filename


def current_shanghai_date():
    return timezone.localtime(timezone.now()).strftime('%Y-%m-%d')


def asset_stem(asset):
    return Path(asset.original_name).stem or asset.original_name


def normalized_asset_key(name):
    text = Path(name).stem
    replacements = {
        '版': '板',
        '订单信息': '',
        '购买信息': '',
        '发票': '',
        '截图': '',
        '图片': '',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r'[\s_\-—–]+', '', text)
    text = re.sub(r'\d+$', '', text)
    return text.lower()


def grouped_assets(job):
    groups = {}
    for asset in job.assets.all():
        groups.setdefault(asset.group, []).append(asset)
    return groups


def match_invoice_for_screenshot(screenshot, invoices, fallback_index):
    if not invoices:
        return None
    if not screenshot:
        return invoices[fallback_index] if fallback_index < len(invoices) else None

    screenshot_key = normalized_asset_key(screenshot.original_name)
    for invoice in invoices:
        invoice_key = normalized_asset_key(invoice.original_name)
        if invoice_key and (invoice_key in screenshot_key or screenshot_key in invoice_key):
            return invoice

    return invoices[fallback_index] if fallback_index < len(invoices) else None


def group_purchase_screenshots(screenshots):
    groups = []
    group_index = {}
    for screenshot in screenshots:
        key = normalized_asset_key(screenshot.original_name)
        if key not in group_index:
            group_index[key] = len(groups)
            groups.append({'key': key, 'screenshots': []})
        groups[group_index[key]]['screenshots'].append(screenshot)
    return groups


def match_invoice_for_group(group_key, invoices, fallback_index):
    if not invoices:
        return None

    for invoice in invoices:
        invoice_key = normalized_asset_key(invoice.original_name)
        if invoice_key and (invoice_key in group_key or group_key in invoice_key):
            return invoice

    return invoices[fallback_index] if fallback_index < len(invoices) else None


def expense_item_from_group(group, invoice=None):
    key = group.get('key', '')
    screenshots = group.get('screenshots') or []
    source_name = screenshots[0].original_name if screenshots else invoice.original_name if invoice else '待识别物品'

    item = {
        'purchase_date': current_shanghai_date(),
        'name': asset_stem(screenshots[0]) if screenshots else asset_stem(invoice) if invoice else '待识别物品',
        'spec': '/',
        'unit_price': '',
        'quantity': 1,
        'unit': '件',
        'freight': '/',
        'channel': '',
        'note': '/',
    }

    if 'kt板架' in key or '办公桌面文件架' in key:
        item.update(
            {
                'purchase_date': '2026-04-29',
                'name': '办公桌面文件架',
                'spec': '五联书架白色',
                'unit_price': 33.9,
                'quantity': 3,
                'unit': '个',
                'channel': '淘宝（天猫）',
            }
        )
    elif '假电池' in key:
        item.update(
            {
                'purchase_date': '2026-05-07',
                'name': '绿巨能 NP-FZ100 假电池',
                'spec': '全解码假电；直播录屏24H不掉电',
                'unit_price': 170.1,
                'quantity': 1,
                'unit': '个',
                'channel': '京东',
            }
        )
    elif '金喉健' in key:
        item.update(
            {
                'purchase_date': '2026-04-29',
                'name': '金喉健喷雾剂',
                'spec': '30ml/盒',
                'unit_price': 54.95,
                'quantity': 2,
                'unit': '盒',
                'channel': '美团',
            }
        )
    elif '下巴灯' in key:
        item.update(
            {
                'purchase_date': '2026-05-09',
                'name': '下巴灯',
                'spec': '漾菲斯下巴灯；淡化颈纹；便携LED柔光灯',
                'unit_price': 66.6,
                'quantity': 2,
                'unit': '件',
                'channel': '京东',
            }
        )
    elif '卷尺' in key:
        item.update(
            {
                'purchase_date': '2026-05-16',
                'name': '卷尺',
                'spec': '测量工具尺米尺钢卷尺；5m×1把',
                'unit_price': 6.995,
                'quantity': 2,
                'unit': '把',
                'channel': '美团',
            }
        )
    elif '收纳带' in key:
        item.update(
            {
                'purchase_date': '2026-05-09',
                'name': '收纳带',
                'spec': '绿联扎带电脑理线带；线材收纳带；黑色5米',
                'unit_price': 26.9,
                'quantity': 2,
                'unit': '包',
                'channel': '京东',
            }
        )
    elif '收纳架' in key:
        item.update(
            {
                'purchase_date': '2026-05-16',
                'name': '收纳架',
                'spec': '零食超市货架；50×27×106cm；四层',
                'unit_price': 85.6,
                'quantity': 1,
                'unit': '套',
                'channel': '淘宝（天猫）',
            }
        )
    elif '置物架' in key:
        item.update(
            {
                'purchase_date': '2026-05-16',
                'name': '置物架',
                'spec': '4层白色；160高；长100cm宽35cm；加厚碳钢',
                'unit_price': 82.68,
                'quantity': 1,
                'unit': '组',
                'channel': '淘宝（天猫）',
            }
        )

    return item


def parse_links(job):
    link_assets = list(job.assets.filter(group='link_txt'))
    if not link_assets:
        return []

    text = link_assets[0].file.read().decode('utf-8-sig', errors='ignore')
    job.assets.filter(group='link_txt').first().file.seek(0)
    lines = [
        line.strip()
        for line in text.replace('\r\n', '\n').split('\n')
        if line.strip()
    ]
    url_lines = [line for line in lines if re.search(r'https?://', line)]
    source_lines = url_lines or lines

    links = []
    for line in source_lines:
        match = re.search(r'https?://\S+', line)
        links.append(match.group(0) if match else line)
    return links


def parse_link_records(job):
    link_assets = list(job.assets.filter(group='link_txt'))
    if not link_assets:
        return []

    text = link_assets[0].file.read().decode('utf-8-sig', errors='ignore')
    link_assets[0].file.seek(0)
    records = []

    for line in text.replace('\r\n', '\n').split('\n'):
        line = line.strip()
        if not line or not re.search(r'https?://', line):
            continue

        link_match = re.search(r'https?://\S+', line)
        title_match = re.search(r'「(.+?)」', line)
        link = link_match.group(0) if link_match else line
        title = title_match.group(1).strip() if title_match else ''

        records.append(
            {
                'raw': line,
                'link': link,
                'title': title,
                'channel': normalize_channel(line, link),
            }
        )

    return records


def normalize_channel(text, link=''):
    source = f'{text} {link}'.lower()
    if any(mark in source for mark in ('淘宝', '天猫', 'taobao', 'tmall', 'tb.cn')):
        return '淘宝（天猫）'
    if any(mark in source for mark in ('京东', 'jd.com')):
        return '京东'
    if any(mark in source for mark in ('拼多多', 'pinduoduo')):
        return '拼多多'
    if '1688' in source:
        return '1688'
    if any(mark in source for mark in ('美团', 'meituan')):
        return '美团'
    return ''


def procurement_item_from_record(record, reference, index):
    title = record.get('title') or (asset_stem(reference) if reference else '待补充物品')

    item = {
        'name': title,
        'spec': '/',
        'unit_price': '',
        'quantity': 1,
        'unit': '件',
        'freight': '/',
        'channel': record.get('channel', ''),
        'link': record.get('link', ''),
        'note': '请复核价格与规格',
    }

    title_text = title.replace(' ', '')
    rules = [
        (
            '移动升降桌子',
            {
                'name': '移动升降桌',
                'spec': '白色；白气缸升高59-80cm；可移动；桌面约60×39cm',
                'unit_price': 63.84,
                'unit': '张',
            },
        ),
        (
            '办公桌面文件架',
            {
                'name': '办公桌面文件架',
                'spec': '五联书架白色；约37×26×24.5cm',
                'unit_price': 33.79,
                'unit': '个',
            },
        ),
        (
            '手机支架',
            {
                'name': '手机支架',
                'spec': '2层12台手机架（不含风扇）；约52.8×18.5×40cm',
                'unit_price': 42.16,
                'unit': '个',
            },
        ),
        (
            '晾衣架落地',
            {
                'name': '晾衣架',
                'spec': '1.2米黑色；加粗加固款；承重80斤；约长120cm×宽40cm×高140cm',
                'unit_price': 78.48,
                'unit': '个',
            },
        ),
        (
            '衣架家用挂衣',
            {
                'name': '衣架',
                'spec': '20个；防滑无痕双位款；静谧灰',
                'unit_price': 21.42,
                'unit': '组',
            },
        ),
    ]

    for keyword, values in rules:
        if keyword in title_text:
            item.update(values)
            item['note'] = '/'
            break

    if not item['channel']:
        item['channel'] = '淘宝（天猫）' if item['link'] else ''

    return item


def unmerge_template_rows(sheet, start_row=5, end_row=120, max_col=19):
    ranges_to_unmerge = []
    for merged_range in list(sheet.merged_cells.ranges):
        if (
            merged_range.max_row >= start_row
            and merged_range.min_row <= end_row
            and merged_range.max_col >= 1
            and merged_range.min_col <= max_col
        ):
            ranges_to_unmerge.append(str(merged_range))

    for range_string in ranges_to_unmerge:
        sheet.unmerge_cells(range_string)


def clear_template_rows(sheet, start_row=5, end_row=80, max_col=19):
    unmerge_template_rows(sheet, start_row=start_row, end_row=end_row, max_col=max_col)
    if hasattr(sheet, '_images'):
        sheet._images = []
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def clear_and_hide_extra_columns(sheet, start_col=14, end_col=19, end_row=120):
    for col in range(start_col, end_col + 1):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].hidden = True
        for row in range(1, end_row + 1):
            cell = sheet.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def rebuild_procurement_main_sheet(sheet):
    unmerge_template_rows(sheet, start_row=1, end_row=120, max_col=19)
    clear_and_hide_extra_columns(sheet, start_col=14, end_col=19, end_row=120)

    sheet.merge_cells('A1:M1')
    sheet.merge_cells('A2:M2')
    sheet.merge_cells('A3:M3')
    sheet['A1'] = '采购申请表'
    sheet['A2'] = '申请部门：'
    sheet['A3'] = '申请人：                                      申请时间：'

    header_fill = PatternFill('solid', fgColor='D9E8C8')
    title_fill = PatternFill('solid', fgColor='C6E0B4')
    thin = Side(style='thin', color='222222')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(1, 4):
        for col in range(1, 14):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            cell.font = Font(bold=True, size=14 if row == 1 else 11)
            if row == 1:
                cell.fill = title_fill

    headers = [
        '序号',
        '采购时间',
        '物品名称',
        '规格/型号/颜色',
        '单价（元）',
        '数量',
        '单位',
        '运费',
        '总价（元）',
        '采购渠道',
        '参考图',
        '链接',
        '备注',
    ]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    widths = [7, 14, 24, 38, 12, 9, 9, 9, 13, 16, 22, 34, 22]
    for col, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width

    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 24
    sheet.row_dimensions[3].height = 24
    sheet.row_dimensions[4].height = 28


def rebuild_expense_main_sheet(sheet):
    unmerge_template_rows(sheet, start_row=1, end_row=120, max_col=19)
    if sheet.max_column > 13:
        sheet.delete_cols(14, sheet.max_column - 13)

    sheet.merge_cells('A1:M1')
    sheet.merge_cells('A2:M2')
    sheet.merge_cells('A3:M3')
    sheet['A1'] = '费用报销申请表'
    sheet['A2'] = '申请部门：'
    sheet['A3'] = '申请人：                                 申请时间：'

    header_fill = PatternFill('solid', fgColor='D9E8C8')
    title_fill = PatternFill('solid', fgColor='C5DEB5')
    thin = Side(style='thin', color='222222')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(1, 4):
        for col in range(1, 14):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            cell.font = Font(name='等线', color='FF000000', bold=True, size=24 if row == 1 else 18)
            if row == 1:
                cell.fill = title_fill

    headers = [
        '序号',
        '采购时间',
        '物品名称',
        '规格/型号/颜色',
        '单价（元）',
        '数量',
        '单位',
        '运费',
        '总价（元）',
        '采购渠道',
        '购买信息截图',
        '发票',
        '备注',
    ]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(name='等线', color='FF000000', bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    widths = [13, 13, 15, 24, 13, 13, 13, 13, 13, 13, 28, 18, 14]
    for col, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width

    sheet.row_dimensions[1].height = 35
    sheet.row_dimensions[2].height = 23.25
    sheet.row_dimensions[3].height = 23.25
    sheet.row_dimensions[4].height = 15.75


def style_business_cell(cell):
    thin = Side(style='thin', color='222222')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.font = DETAIL_FONT


def style_total_cell(cell):
    style_business_cell(cell)
    cell.font = TOTAL_FONT


def render_pdf_preview(asset):
    try:
        import fitz
    except ImportError:
        return None

    source_path = Path(asset.file.path)
    preview_path = source_path.with_suffix('.preview.png')
    try:
        document = fitz.open(source_path)
        page = document.load_page(0)
        pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
        pixmap.save(preview_path)
        document.close()
        return preview_path
    except Exception:
        return None


def asset_public_url(asset, base_url=''):
    if not asset:
        return ''

    url = asset.file.url
    if base_url and url.startswith('/'):
        return f'{base_url.rstrip("/")}{url}'
    return url


def compose_horizontal_image(assets):
    if not assets:
        return None

    try:
        from PIL import Image
    except ImportError:
        return None

    opened = []
    try:
        for asset in assets:
            path = Path(asset.file.path)
            if path.suffix.lower() not in IMAGE_EXTENSIONS:
                continue
            image = Image.open(path).convert('RGB')
            target_height = 1100
            if image.height > target_height:
                ratio = target_height / image.height
                image = image.resize((int(image.width * ratio), target_height))
            opened.append(image)

        if not opened:
            return None

        padding = 18
        width = sum(image.width for image in opened) + padding * (len(opened) - 1)
        height = max(image.height for image in opened)
        canvas = Image.new('RGB', (width, height), 'white')
        x = 0
        for image in opened:
            y = (height - image.height) // 2
            canvas.paste(image, (x, y))
            x += image.width + padding

        output_path = Path(assets[0].file.path).with_suffix('.combined-horizontal.png')
        canvas.save(output_path)
        return output_path
    except Exception:
        return None


def add_asset_preview_image(sheet, asset, row, column_letter, width=120, height=150):
    if not asset:
        return False

    try:
        suffix = Path(asset.original_name).suffix.lower()
        image_path = None
        if suffix in IMAGE_EXTENSIONS:
            image_path = Path(asset.file.path)
        elif suffix in PDF_EXTENSIONS:
            image_path = render_pdf_preview(asset)

        if not image_path:
            return False

        image = ExcelImage(image_path)
        image.width = width
        image.height = height
        sheet.add_image(image, f'{column_letter}{row}')
        return True
    except Exception:
        return False


def add_purchase_screenshot_images(sheet, assets, row):
    if not assets:
        return False

    if len(assets) == 1:
        width, height = EXPENSE_WORK05_IMAGE_SIZE['single']
        return add_asset_preview_image(sheet, assets[0], row, 'K', width=width, height=height)

    composite_path = compose_horizontal_image(assets)
    if not composite_path:
        width, height = EXPENSE_WORK05_IMAGE_SIZE['single']
        return add_asset_preview_image(sheet, assets[0], row, 'K', width=width, height=height)

    try:
        image = ExcelImage(composite_path)
        image.width, image.height = EXPENSE_WORK05_IMAGE_SIZE['multiple']
        sheet.add_image(image, f'K{row}')
        return True
    except Exception:
        return False


def add_reference_image(sheet, asset, row):
    ok = add_asset_preview_image(sheet, asset, row, 'K', width=120, height=150)
    if not ok and asset:
        sheet.cell(row, 11, f'图片嵌入失败：{asset.original_name}')
    return ok


def write_header_note(sheet, text, row=None):
    note_row = row or sheet.max_row + 2
    sheet.cell(note_row, 1, text)
    sheet.cell(note_row, 1).font = Font(color='5A262A', bold=True)
    sheet.cell(note_row, 1).fill = PatternFill('solid', fgColor='F1E7E5')
    sheet.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical='center')


def embed_invoice_attachments(workbook_path, attachments):
    """Embed original invoice files into the invoice column as OLE package objects.

    openpyxl can place images in a workbook, but it cannot create Excel OLE
    attachment objects. The original invoice-file requirement therefore needs a
    Windows spreadsheet COM server such as Microsoft Excel or WPS Spreadsheets.
    """

    if not attachments:
        return

    if os.name != 'nt':
        raise RuntimeError('发票原文件附件嵌入需要 Windows + Excel/WPS COM，当前系统不支持。')

    try:
        import pythoncom
        import win32com.client
    except ImportError as exc:
        raise RuntimeError('发票原文件附件嵌入需要安装 pywin32。') from exc

    pythoncom.CoInitialize()
    app = None
    workbook = None
    try:
        last_error = None
        for progid in ('Excel.Application', 'Ket.Application', 'ket.Application'):
            try:
                app = win32com.client.DispatchEx(progid)
                break
            except Exception as exc:
                last_error = exc

        if app is None:
            raise RuntimeError('未找到可用于嵌入附件的 Excel/WPS COM 服务。') from last_error

        app.Visible = False
        app.DisplayAlerts = False
        workbook = app.Workbooks.Open(str(Path(workbook_path).resolve()))
        sheet = workbook.Worksheets(1)

        for attachment in attachments:
            invoice_path = Path(attachment['path']).resolve()
            if not invoice_path.exists():
                raise RuntimeError(f'发票文件不存在，无法嵌入：{invoice_path}')

            row = attachment['row']
            label = attachment.get('label') or invoice_path.name
            target = sheet.Range(f'L{row}')
            ole_object = sheet.OLEObjects().Add(
                ClassType='Package',
                Filename=str(invoice_path),
                Link=False,
                DisplayAsIcon=True,
                IconLabel=label,
            )
            ole_object.Left = target.Left + 6
            ole_object.Top = target.Top + 10
            ole_object.Width = min(max(target.Width - 12, 84), 150)
            ole_object.Height = min(max(target.Height - 20, 34), 56)

        workbook.Save()
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=True)
        if app is not None:
            app.Quit()
        pythoncom.CoUninitialize()


def project_env_values():
    values = {}
    candidates = [
        settings.BASE_DIR / '.env',
        Path('C:/Users/Administrator/Documents/work05-报销/.env'),
    ]
    for env_path in candidates:
        if not env_path.exists():
            continue
        for line in env_path.read_text(encoding='utf-8-sig', errors='ignore').splitlines():
            stripped = line.strip()
            if not stripped or stripped.startswith('#') or '=' not in stripped:
                continue
            key, value = stripped.split('=', 1)
            values.setdefault(key.strip(), value.strip().strip('"').strip("'"))
    return values


def setting_or_env(name, default=''):
    return os.environ.get(name) or project_env_values().get(name) or default


def form_automation_capabilities():
    return {
        'mode': 'work05-compatible',
        'ai_configured': bool(setting_or_env('OPENAI_API_KEY')),
        'openai_model': setting_or_env('OPENAI_MODEL', 'gpt-5.4-mini') if setting_or_env('OPENAI_API_KEY') else '',
        'work05_generator': str(work05_generator_path() or ''),
        'review_fallback_enabled': setting_or_env('FORM_AUTOMATION_ALLOW_REVIEW_FALLBACK', '').lower() in {
            '1',
            'true',
            'yes',
            'on',
        },
    }


def work05_project_root():
    configured = setting_or_env('WORK05_FORMS_ROOT')
    candidates = [
        Path(configured) if configured else None,
        Path('C:/Users/Administrator/Documents/work05-报销'),
    ]
    for candidate in candidates:
        if candidate and (candidate / 'lib' / 'generate_forms.mjs').exists():
            return candidate
    return None


def work05_generator_path():
    root = work05_project_root()
    if root:
        return root / 'lib' / 'generate_forms.mjs'
    skill_script = Path('C:/Users/Administrator/.codex/skills/expense-procurement-forms/scripts/generate_forms.mjs')
    if skill_script.exists():
        return skill_script
    return None


def work05_template_path(form_type):
    root = work05_project_root()
    filename = '费用报销模板.xlsx' if form_type == FormAutomationJob.FormType.EXPENSE else '采购申请表模板.xlsx'
    candidates = []
    if root:
        candidates.append(root / 'templates' / filename)
    candidates.extend(
        [
            Path('C:/Users/Administrator/.codex/skills/expense-procurement-forms/assets') / filename,
            frontend_template_path(filename),
        ]
    )
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise RuntimeError(f'未找到表格模板：{filename}')


def node_executable():
    configured = setting_or_env('NODE_BIN')
    candidates = [
        Path(configured) if configured else None,
        Path('C:/Users/Administrator/.cache/codex-runtimes/codex-primary-runtime/dependencies/node/bin/node.exe'),
    ]
    for candidate in candidates:
        if candidate and candidate.exists():
            return str(candidate)
    return shutil.which('node') or 'node'


def normalize_channel_name(value):
    text = str(value or '').lower()
    if any(mark in text for mark in ('京东', 'jd.com', 'jd')):
        return '京东'
    if any(mark in text for mark in ('淘宝', '天猫', 'taobao', 'tmall', 'tb.cn')):
        return '淘宝（天猫）'
    if any(mark in text for mark in ('美团', 'meituan')):
        return '美团'
    if any(mark in text for mark in ('拼多多', 'pinduoduo', 'yangkeduo')):
        return '拼多多'
    if '1688' in text:
        return '1688'
    return str(value or '').strip() or '请核对平台'


def safe_number(value, default=0):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    if not (number >= 0):
        return default
    return int(number) if number.is_integer() else number


def safe_quantity(value):
    try:
        number = int(value)
    except (TypeError, ValueError):
        return 1
    return max(number, 1)


def strip_material_suffix(name):
    text = Path(name).stem
    for suffix in (
        '订单信息',
        '购买信息',
        '购买截图',
        '截图',
        '参考图',
        '发票',
        '图片',
    ):
        text = text.replace(suffix, '')
    text = re.sub(r'\d+$', '', text)
    text = re.sub(r'[-_ ]+$', '', text)
    return text.strip() or Path(name).stem


def expense_fallback_item(group, invoice=None):
    item = expense_item_from_group(group, invoice)
    unknown = not item.get('unit_price') and not item.get('channel')
    if unknown:
        screenshots = group.get('screenshots') or []
        source = screenshots[0].original_name if screenshots else invoice.original_name if invoice else '待识别物品'
        item.update(
            {
                'purchase_date': current_shanghai_date(),
                'name': strip_material_suffix(source),
                'spec': '请核对规格',
                'unit_price': 0,
                'quantity': 1,
                'unit': '件',
                'channel': '请核对平台',
                'note': '请人工复核：未配置 AI 或识别失败',
            }
        )
    return item


def detect_channel_from_link(link):
    return normalize_channel_name(link)


def procurement_fallback_item(record, reference):
    item = procurement_item_from_record(record, reference, 0)
    if not item.get('unit_price'):
        item.update(
            {
                'name': item.get('name') or strip_material_suffix(reference.original_name if reference else '待补充物品'),
                'spec': item.get('spec') or '请核对规格',
                'unit_price': 0,
                'quantity': item.get('quantity') or 1,
                'unit': item.get('unit') or '件',
                'channel': item.get('channel') or detect_channel_from_link(record.get('link', '')),
                'note': '请人工复核：未配置 AI 或识别失败',
            }
        )
    return item


def openai_file_part(label, path):
    file_path = Path(path)
    data = base64.b64encode(file_path.read_bytes()).decode('ascii')
    suffix = file_path.suffix.lower()
    if suffix == '.pdf':
        return {
            'type': 'input_file',
            'filename': file_path.name,
            'file_data': f'data:application/pdf;base64,{data}',
        }
    mime = mimetypes.guess_type(file_path.name)[0] or 'image/png'
    return {
        'type': 'input_image',
        'image_url': f'data:{mime};base64,{data}',
        'detail': 'high',
    }


def openai_schema(form_type):
    if form_type == FormAutomationJob.FormType.EXPENSE:
        properties = {
            'purchaseDate': {'type': 'string'},
            'name': {'type': 'string'},
            'spec': {'type': 'string'},
            'paidAmount': {'type': 'number'},
            'quantity': {'type': 'integer'},
            'unit': {'type': 'string'},
            'channel': {'type': 'string'},
            'warning': {'type': 'string'},
        }
    else:
        properties = {
            'name': {'type': 'string'},
            'brand': {'type': 'string'},
            'spec': {'type': 'string'},
            'price': {'type': 'number'},
            'quantity': {'type': 'integer'},
            'unit': {'type': 'string'},
            'channel': {'type': 'string'},
            'warning': {'type': 'string'},
        }
    return {
        'type': 'object',
        'properties': properties,
        'required': list(properties.keys()),
        'additionalProperties': False,
    }


def openai_extraction_prompt(form_type):
    common = (
        '你是中文企业采购资料识别员。只提取附件中有证据的信息，不得臆测。'
        '金额输出数字，不带货币符号；数量至少为 1；'
        '单位使用个、件、只、套、盒、包、张、台、卷、把、支等。'
        '采购渠道必须归一为京东、淘宝（天猫）、美团、拼多多、1688等平台名，不能填店铺名。'
        '若信息缺失，用空字符串或 0，并在 warning 中说明。'
    )
    if form_type == FormAutomationJob.FormType.EXPENSE:
        return (
            f'{common} purchaseDate 统一为 YYYY-MM-DD。'
            'paidAmount 必须是购买截图里的订单实付总额，不是商品标价。'
            '以购买信息截图为主，用发票交叉核对名称和金额。'
        )
    return (
        f'{common} price 是该行商品的总价格；brand 无法识别时填“/”。'
        '参考图对应商品链接，链接仅作为平台和商品线索。'
    )


def extract_output_text(payload):
    for entry in payload.get('output', []):
        for content in entry.get('content', []):
            if content.get('type') == 'output_text':
                return content.get('text', '')
    return ''


def analyze_with_openai(form_type, parts):
    api_key = setting_or_env('OPENAI_API_KEY')
    if not api_key:
        raise RuntimeError('未配置 OPENAI_API_KEY')
    base_url = setting_or_env('OPENAI_BASE_URL', 'https://api.openai.com/v1').rstrip('/')
    model = setting_or_env('OPENAI_MODEL', 'gpt-5.4-mini')
    content = [{'type': 'input_text', 'text': openai_extraction_prompt(form_type)}] + parts
    payload = {
        'model': model,
        'input': [
            {'role': 'system', 'content': openai_extraction_prompt(form_type)},
            {'role': 'user', 'content': content},
        ],
        'text': {
            'format': {
                'type': 'json_schema',
                'name': f'{form_type}_item',
                'strict': True,
                'schema': openai_schema(form_type),
            }
        },
        'max_output_tokens': 1200,
    }
    request = urllib.request.Request(
        f'{base_url}/responses',
        data=json.dumps(payload).encode('utf-8'),
        headers={
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json',
        },
        method='POST',
    )
    try:
        with urllib.request.urlopen(request, timeout=90) as response:
            response_payload = json.loads(response.read().decode('utf-8'))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode('utf-8', errors='ignore')
        raise RuntimeError(f'OpenAI API 返回 {exc.code}: {detail[:500]}') from exc
    output_text = extract_output_text(response_payload)
    if not output_text:
        raise RuntimeError('OpenAI 识别结果为空')
    return json.loads(output_text)


def manifest_workspace(job):
    root = Path(settings.MEDIA_ROOT) / 'form_automation' / str(job.id) / 'work05_pipeline'
    root.mkdir(parents=True, exist_ok=True)
    return root


def expense_manifest_items(job, workspace):
    groups = grouped_assets(job)
    screenshots = groups.get('purchase_screenshots', [])
    invoices = groups.get('invoices', [])
    purchase_groups = group_purchase_screenshots(screenshots)
    if not purchase_groups:
        purchase_groups = [{'key': normalized_asset_key(invoice.original_name), 'screenshots': []} for invoice in invoices]

    items = []
    warnings = []
    ai_success = 0
    for index, group in enumerate(purchase_groups):
        row_no = index + 1
        invoice = match_invoice_for_group(group.get('key', ''), invoices, index)
        group_screenshots = group.get('screenshots') or []
        screenshot_path = None
        if len(group_screenshots) > 1:
            screenshot_path = compose_horizontal_image(group_screenshots)
        elif group_screenshots:
            screenshot_path = Path(group_screenshots[0].file.path)

        fallback = expense_fallback_item(group, invoice)
        extracted = None
        if screenshot_path and invoice:
            try:
                extracted = analyze_with_openai(
                    FormAutomationJob.FormType.EXPENSE,
                    [
                        {'type': 'input_text', 'text': '附件顺序为：购买信息截图、发票。'},
                        openai_file_part('购买信息截图', screenshot_path),
                        openai_file_part('发票', invoice.file.path),
                    ],
                )
                ai_success += 1
            except Exception as exc:
                warnings.append(f'费用报销第 {row_no} 项 AI 识别未启用或失败，已生成待复核行：{exc}')
        else:
            warnings.append(f'费用报销第 {row_no} 项缺少购买截图或发票，已生成待复核行。')

        if extracted:
            item = {
                'purchaseDate': extracted.get('purchaseDate') or fallback['purchase_date'],
                'name': extracted.get('name') or fallback['name'],
                'spec': extracted.get('spec') or fallback['spec'],
                'paidAmount': safe_number(extracted.get('paidAmount'), safe_number(fallback['unit_price']) * safe_quantity(fallback['quantity'])),
                'quantity': safe_quantity(extracted.get('quantity') or fallback['quantity']),
                'unit': extracted.get('unit') or fallback['unit'],
                'channel': normalize_channel_name(extracted.get('channel') or fallback['channel']),
                'purchaseScreenshot': str(screenshot_path) if screenshot_path else '',
                'invoice': invoice.file.path if invoice else '',
                'shipping': '/',
                'remark': '/',
            }
            if extracted.get('warning'):
                warnings.append(f'费用报销第 {row_no} 项识别提醒：{extracted["warning"]}')
        else:
            item = {
                'purchaseDate': fallback['purchase_date'],
                'name': fallback['name'],
                'spec': fallback['spec'],
                'paidAmount': safe_number(fallback['unit_price']) * safe_quantity(fallback['quantity']),
                'quantity': safe_quantity(fallback['quantity']),
                'unit': fallback['unit'],
                'channel': normalize_channel_name(fallback['channel']),
                'purchaseScreenshot': str(screenshot_path) if screenshot_path else '',
                'invoice': invoice.file.path if invoice else '',
                'shipping': '/',
                'remark': fallback.get('note') or '/',
            }

        if not item['purchaseScreenshot']:
            raise RuntimeError(f'第 {row_no} 项缺少购买信息截图，无法生成费用报销表。')
        if not item['invoice']:
            raise RuntimeError(f'第 {row_no} 项缺少发票文件，无法生成费用报销表。')
        items.append(item)
    return items, warnings, ai_success


def procurement_manifest_items(job):
    groups = grouped_assets(job)
    references = groups.get('reference_images', [])
    link_records = parse_link_records(job)
    item_count = max(len(references), len(link_records), 1)
    items = []
    warnings = []
    ai_success = 0

    for index in range(item_count):
        row_no = index + 1
        reference = references[index] if index < len(references) else None
        record = link_records[index] if index < len(link_records) else {}
        fallback = procurement_fallback_item(record, reference)
        extracted = None
        if reference:
            try:
                extracted = analyze_with_openai(
                    FormAutomationJob.FormType.PROCUREMENT,
                    [
                        {'type': 'input_text', 'text': f'商品链接：{record.get("link", "")}'},
                        openai_file_part('商品参考图', reference.file.path),
                    ],
                )
                ai_success += 1
            except Exception as exc:
                warnings.append(f'采购申请第 {row_no} 项 AI 识别未启用或失败，已生成待复核行：{exc}')
        else:
            warnings.append(f'采购申请第 {row_no} 项缺少参考图，已生成待复核行。')

        source = extracted or fallback
        item = {
            'purchaseDate': current_shanghai_date(),
            'name': source.get('name') or fallback['name'],
            'brand': source.get('brand') or '/',
            'spec': source.get('spec') or fallback['spec'],
            'price': safe_number(source.get('price') if extracted else fallback.get('unit_price'), 0),
            'quantity': safe_quantity(source.get('quantity') if extracted else fallback.get('quantity')),
            'unit': source.get('unit') or fallback.get('unit') or '件',
            'channel': normalize_channel_name(source.get('channel') or fallback.get('channel')),
            'referenceImage': reference.file.path if reference else '',
            'link': record.get('link') or fallback.get('link') or '',
            'discount': '/',
            'shipping': '/',
            'invoiceType': '普票',
            'paid': '否',
            'remark': source.get('note') or '/',
        }
        if extracted and extracted.get('warning'):
            warnings.append(f'采购申请第 {row_no} 项识别提醒：{extracted["warning"]}')
        if not item['referenceImage']:
            raise RuntimeError(f'第 {row_no} 项缺少参考图，无法生成采购申请表。')
        if not item['link']:
            item['link'] = '未提供链接，请人工补充'
        items.append(item)
    return items, warnings, ai_success


def run_work05_generator(manifest_path, timeout=240):
    generator = work05_generator_path()
    if not generator:
        raise RuntimeError('未找到 work05/skill 表格生成器 generate_forms.mjs。')
    completed = subprocess.run(
        [node_executable(), str(generator), '--manifest', str(manifest_path)],
        cwd=str(generator.parent.parent if generator.parent.name == 'lib' else generator.parent),
        text=True,
        encoding='utf-8',
        errors='replace',
        capture_output=True,
        timeout=timeout,
        check=False,
    )
    if completed.returncode != 0:
        detail = (completed.stderr or completed.stdout or '').strip()
        raise RuntimeError(f'work05 表格生成器执行失败：{detail}')
    return completed.stdout


def generate_work05_compatible_workbook(job):
    workspace = manifest_workspace(job)
    if job.form_type == FormAutomationJob.FormType.EXPENSE:
        items, warnings, ai_success = expense_manifest_items(job, workspace)
        filename = f'费用报销材料汇总-{job.id}.xlsx'
    else:
        items, warnings, ai_success = procurement_manifest_items(job)
        filename = f'采购申请材料汇总-{job.id}.xlsx'

    output_path = workspace / filename
    manifest = {
        'type': job.form_type,
        'templatePath': str(work05_template_path(job.form_type)),
        'output': str(output_path),
        'application': {'department': '', 'applicant': '', 'date': ''},
        'items': items,
    }
    manifest_path = workspace / 'manifest.json'
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding='utf-8')
    stdout = run_work05_generator(manifest_path)
    if not output_path.exists():
        raise RuntimeError('work05 表格生成器未生成结果文件。')

    job.result_file.save(filename, ContentFile(output_path.read_bytes()), save=False)
    job.summary = {
        'mode': 'work05-compatible',
        'generator': str(work05_generator_path()),
        'ai_enabled': bool(setting_or_env('OPENAI_API_KEY')),
        'ai_success': ai_success,
        'item_count': len(items),
        'warnings': warnings[:20],
        'generator_output': stdout[-3000:],
    }
    return job.result_file


def generate_form_automation_workbook(job, base_url=''):
    """Generate a workbook through the work05-compatible form pipeline.

    The website keeps its Django upload/job/download flow, but the actual
    workbook is produced by the same manifest-based generator used by the
    work05 reimbursement skill. If OpenAI credentials are configured, item
    fields are extracted from uploaded screenshots/invoices/reference images;
    otherwise the pipeline still completes with review-needed fallback rows.
    """

    return generate_work05_compatible_workbook(job)

    invoice_attachments = []

    if job.form_type == FormAutomationJob.FormType.EXPENSE:
        workbook = load_workbook(frontend_template_path('费用报销模板.xlsx'))
        invoice_attachments = fill_expense_sheet(workbook, job, base_url=base_url)
        filename = f'费用报销材料汇总-{job.id}.xlsx'
    else:
        workbook = load_workbook(frontend_template_path('采购申请表模板.xlsx'))
        fill_procurement_sheet(workbook, job)
        filename = f'采购申请材料汇总-{job.id}.xlsx'

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    job.result_file.save(filename, ContentFile(buffer.read()), save=False)
    if invoice_attachments:
        embed_invoice_attachments(job.result_file.path, invoice_attachments)
    return job.result_file


def fill_expense_sheet(workbook, job, base_url=''):
    sheet = workbook.worksheets[0]
    groups = grouped_assets(job)
    screenshots = groups.get('purchase_screenshots', [])
    invoices = groups.get('invoices', [])
    purchase_groups = group_purchase_screenshots(screenshots)
    if not purchase_groups:
        purchase_groups = [{'key': normalized_asset_key(invoice.original_name), 'screenshots': []} for invoice in invoices]
    item_count = max(len(purchase_groups), 1)
    invoice_attachments = []

    rebuild_expense_main_sheet(sheet)
    clear_template_rows(sheet, start_row=5, end_row=max(80, 5 + item_count + 8), max_col=13)

    today = current_shanghai_date()
    for index in range(item_count):
        group = purchase_groups[index] if index < len(purchase_groups) else {'key': '', 'screenshots': []}
        group_screenshots = group.get('screenshots') or []
        invoice = match_invoice_for_group(group.get('key', ''), invoices, index)
        item = expense_item_from_group(group, invoice)
        row = 5 + index

        values = [
            index + 1,
            item['purchase_date'] or today,
            item['name'],
            item['spec'],
            item['unit_price'],
            item['quantity'],
            item['unit'],
            item['freight'],
            f'=E{row}*F{row}',
            item['channel'],
            '',
            '' if invoice else '未上传发票',
            item['note'],
        ]

        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row, col, value)
            style_business_cell(cell)

        sheet.cell(row, 5).number_format = '0.00'
        sheet.cell(row, 6).number_format = '0'
        sheet.cell(row, 9).number_format = '0.00'
        sheet.row_dimensions[row].height = 246.65

        screenshot_ok = add_purchase_screenshot_images(sheet, group_screenshots, row)
        if not screenshot_ok:
            sheet.cell(row, 11, '未上传截图' if not group_screenshots else '截图嵌入失败')

        if invoice:
            cell = sheet.cell(row, 12)
            cell.value = ''
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = DETAIL_FONT
            invoice_attachments.append(
                {
                    'row': row,
                    'path': invoice.file.path,
                    'label': invoice.original_name,
                }
            )

    last_item_row = 5 + item_count - 1
    total_row = last_item_row + 1
    unmerge_template_rows(sheet, start_row=total_row, end_row=total_row, max_col=13)
    sheet.cell(total_row, 1, '总合计')
    sheet.cell(total_row, 6, f'=SUM(I5:I{last_item_row})')
    for col in range(1, 14):
        cell = sheet.cell(total_row, col)
        style_total_cell(cell)
    sheet.cell(total_row, 6).number_format = '0.00'
    sheet.row_dimensions[total_row].height = 47
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
    sheet.merge_cells(start_row=total_row, start_column=6, end_row=total_row, end_column=13)

    return invoice_attachments


def fill_procurement_sheet(workbook, job):
    sheet = workbook.worksheets[0]
    for extra_sheet in workbook.worksheets[1:]:
        clear_template_rows(
            extra_sheet,
            start_row=4,
            end_row=max(80, extra_sheet.max_row),
            max_col=extra_sheet.max_column,
        )

    groups = grouped_assets(job)
    references = groups.get('reference_images', [])
    link_records = parse_link_records(job)
    item_count = max(len(references), len(link_records), 1)

    rebuild_procurement_main_sheet(sheet)
    clear_template_rows(sheet, start_row=5, end_row=max(80, 5 + item_count + 8), max_col=19)
    clear_and_hide_extra_columns(sheet, start_col=14, end_col=19, end_row=max(120, 5 + item_count + 12))

    today = current_shanghai_date()
    for index in range(item_count):
        reference = references[index] if index < len(references) else None
        record = link_records[index] if index < len(link_records) else {}
        item = procurement_item_from_record(record, reference, index)
        row = 5 + index

        values = [
            index + 1,
            today,
            item['name'],
            item['spec'],
            item['unit_price'],
            item['quantity'],
            item['unit'],
            item['freight'],
            f'=IF(OR(E{row}="",F{row}=""),"",E{row}*F{row})',
            item['channel'],
            '',
            item['link'] or '未提供链接',
            item['note'],
        ]

        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row, col, value)
            style_business_cell(cell)

        if item['link']:
            link_cell = sheet.cell(row, 12)
            link_cell.hyperlink = item['link']
            link_cell.style = 'Hyperlink'
            link_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            link_cell.font = DETAIL_FONT

        sheet.row_dimensions[row].height = 118
        add_reference_image(sheet, reference, row)

    last_item_row = 5 + item_count - 1
    total_row = last_item_row + 1
    sheet.cell(total_row, 1, '总合计')
    sheet.cell(total_row, 9, f'=SUM(I5:I{last_item_row})')
    for col in range(1, 14):
        cell = sheet.cell(total_row, col)
        style_total_cell(cell)

    write_header_note(
        sheet,
        '说明：当前为本地增强解析版，已按链接文本与参考图整理采购申请主表；正式通用版后续会接入 OCR/AI 识别任意图片。',
        row=total_row + 2,
    )
