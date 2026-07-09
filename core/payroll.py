from io import BytesIO
from pathlib import Path

from django.core.files.base import ContentFile
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


FILE_LABELS = (
    ('host_schedule', '主播排班表'),
    ('controller_schedule', '场控排班表'),
    ('trial_schedule', '试播间排班表'),
    ('host_data', '主播数据'),
)


def _file_summary(job, field_name):
    file_field = getattr(job, field_name)
    return {
        'name': file_field.name.split('/')[-1] if file_field else '',
        'size': file_field.size if file_field else 0,
        'path': file_field.path if file_field else '',
    }


def _payroll_period(job):
    if job.week_start and job.week_end:
        return f'{job.week_start.isoformat()} - {job.week_end.isoformat()}'
    if job.week_start:
        return f'{job.week_start.isoformat()} - 未填写'
    if job.week_end:
        return f'未填写 - {job.week_end.isoformat()}'
    return '未填写'


def _style_title(cell):
    cell.fill = PatternFill('solid', fgColor='4C2528')
    cell.font = Font(color='FFFFFF', bold=True, size=14)
    cell.alignment = Alignment(horizontal='center', vertical='center')


def _style_header(cell):
    cell.fill = PatternFill('solid', fgColor='F1E7E5')
    cell.font = Font(color='4C2528', bold=True)
    cell.alignment = Alignment(vertical='center', wrap_text=True)


def generate_placeholder_payroll(job):
    """Generate the local second-step payroll preparation workbook.

    The web flow now receives the actual uploaded files, records the selected
    room and payroll period, reads the host-data workbook when possible, and
    returns a downloadable workbook. The screenshot-to-schedule recognition and
    final rule-driven payroll engine can be plugged into this same endpoint.
    """

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = '薪资计算准备包'

    summary_sheet.merge_cells('A1:D1')
    summary_sheet['A1'] = '造物者兼职薪资计算 - 本地准备包'
    _style_title(summary_sheet['A1'])
    summary_sheet.row_dimensions[1].height = 28

    metadata = [
        ('任务编号', str(job.id)),
        ('选择直播间', job.room_type),
        ('薪资计算周期', _payroll_period(job)),
        ('生成时间', timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M:%S')),
        ('当前阶段', '第二步：已完成上传文件读取、任务处理和下载闭环；待接入截图识别与正式薪资规则引擎。'),
    ]

    row = 3
    for label, value in metadata:
        summary_sheet.cell(row=row, column=1, value=label)
        summary_sheet.cell(row=row, column=2, value=value)
        summary_sheet.cell(row=row, column=1).font = Font(color='4C2528', bold=True)
        row += 1

    row += 1
    summary_sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    summary_sheet.cell(row=row, column=1, value='已上传文件')
    _style_header(summary_sheet.cell(row=row, column=1))
    row += 1

    headers = ['文件类型', '文件名', '文件大小 KB', '处理状态']
    for col, header in enumerate(headers, start=1):
        _style_header(summary_sheet.cell(row=row, column=col, value=header))
    row += 1

    status_notes = {
        'host_schedule': '已保存；下一步由截图识别转换为排班 JSON',
        'controller_schedule': '已保存；下一步由截图识别转换为场控班次',
        'trial_schedule': '已保存；下一步由截图识别转换为试播/彩排班次',
        'host_data': '已保存；本文件会尝试读取并生成预览',
    }

    for field_name, label in FILE_LABELS:
        summary = _file_summary(job, field_name)
        summary_sheet.append([
            label,
            summary['name'],
            round(summary['size'] / 1024, 2),
            status_notes[field_name],
        ])

    row = summary_sheet.max_row + 2
    summary_sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    summary_sheet.cell(
        row=row,
        column=1,
        value='说明：正式薪资计算还需要把三张排班截图转成结构化排班 JSON。当前本地第二步已经把网站后端处理链路打通，后续可直接替换为 live-payroll 规则引擎。',
    )
    summary_sheet.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    summary_sheet.row_dimensions[row].height = 54

    add_host_data_preview(workbook, job)

    widths = [22, 44, 16, 64]
    for sheet in workbook.worksheets:
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        for row_cells in sheet.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(
                    vertical='center',
                    wrap_text=True,
                    horizontal=cell.alignment.horizontal,
                )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f'creator-payroll-preparation-{job.id}.xlsx'
    job.result_file.save(filename, ContentFile(buffer.read()), save=False)
    return job.result_file


def add_host_data_preview(workbook, job):
    host_data = _file_summary(job, 'host_data')
    preview_sheet = workbook.create_sheet('主播数据预览')

    preview_sheet.merge_cells('A1:E1')
    preview_sheet['A1'] = '主播数据预览'
    _style_title(preview_sheet['A1'])

    if not host_data['path'] or Path(host_data['path']).suffix.lower() not in {'.xlsx', '.xlsm'}:
        preview_sheet['A3'] = '主播数据不是可预览的 Excel 文件，已保存原文件，后续计算时仍可读取。'
        return

    try:
        source_workbook = load_workbook(host_data['path'], data_only=True, read_only=True)
        source_sheet = source_workbook.worksheets[0]
        for row_index, source_row in enumerate(
            source_sheet.iter_rows(min_row=1, max_row=12, values_only=True),
            start=3,
        ):
            for col_index, value in enumerate(source_row[:8], start=1):
                preview_sheet.cell(row=row_index, column=col_index, value=value)
        source_workbook.close()
    except Exception as exc:
        preview_sheet['A3'] = f'主播数据预览失败，但原文件已保存：{exc}'
